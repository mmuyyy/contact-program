from flask import request, jsonify, send_file
import mysql.connector
from .models import get_db_connection
import openpyxl
from io import BytesIO


def register_routes(app):
    # 1. 书签/收藏联系人
    @app.route('/api/contacts/<int:contact_id>/favorite', methods=['PUT'])
    def toggle_favorite(contact_id):
        data = request.json
        is_favorite = data.get('is_favorite')
        connection = get_db_connection(app.config['db_config'])  # 假设在 app.py 中设置 app.config['db_config'] = db_config
        if connection:
            cursor = connection.cursor()
            cursor.execute('UPDATE contacts SET is_favorite = %s WHERE id = %s', (is_favorite, contact_id))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': '收藏状态已更新'}), 200
        return jsonify({'error': '数据库连接失败'}), 500

    # 2. 添加多个联系方式（添加/编辑联系人的部分）
    @app.route('/api/contacts', methods=['POST'])
    def add_contact():
        data = request.json
        name = data.get('name')
        methods = data.get('methods', [])  # 列表，如 {'type': 'phone', 'value': '123456'}
        connection = get_db_connection(app.config['db_config'])
        if connection:
            cursor = connection.cursor()
            cursor.execute('INSERT INTO contacts (name) VALUES (%s)', (name,))
            contact_id = cursor.lastrowid
            for method in methods:
                cursor.execute('INSERT INTO contact_methods (contact_id, type, value) VALUES (%s, %s, %s)',
                               (contact_id, method['type'], method['value']))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': '联系人已添加', 'id': contact_id}), 201
        return jsonify({'error': '数据库连接失败'}), 500

    @app.route('/api/contacts/<int:contact_id>', methods=['PUT'])
    def update_contact(contact_id):
        data = request.json
        name = data.get('name')
        methods = data.get('methods', [])  # 替换现有联系方式
        connection = get_db_connection(app.config['db_config'])
        if connection:
            cursor = connection.cursor()
            if name:
                cursor.execute('UPDATE contacts SET name = %s WHERE id = %s', (name, contact_id))
            # 删除现有联系方式
            cursor.execute('DELETE FROM contact_methods WHERE contact_id = %s', (contact_id,))
            # 添加新联系方式
            for method in methods:
                cursor.execute('INSERT INTO contact_methods (contact_id, type, value) VALUES (%s, %s, %s)',
                               (contact_id, method['type'], method['value']))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': '联系人已更新'}), 200
        return jsonify({'error': '数据库连接失败'}), 500

    # 获取所有联系人
    @app.route('/api/contacts', methods=['GET'])
    def get_contacts():
        connection = get_db_connection(app.config['db_config'])
        if connection:
            cursor = connection.cursor(dictionary=True)
            cursor.execute('SELECT * FROM contacts')
            contacts = cursor.fetchall()
            for contact in contacts:
                cursor.execute('SELECT type, value FROM contact_methods WHERE contact_id = %s', (contact['id'],))
                contact['methods'] = cursor.fetchall()
            cursor.close()
            connection.close()
            return jsonify(contacts), 200
        return jsonify({'error': '数据库连接失败'}), 500

    # 删除联系人
    @app.route('/api/contacts/<int:contact_id>', methods=['DELETE'])
    def delete_contact(contact_id):
        connection = get_db_connection(app.config['db_config'])
        if connection:
            cursor = connection.cursor()
            cursor.execute('DELETE FROM contacts WHERE id = %s', (contact_id,))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': '联系人已删除'}), 200
        return jsonify({'error': '数据库连接失败'}), 500

    # 3. 导入和导出
    @app.route('/api/export', methods=['GET'])
    def export_contacts():
        connection = get_db_connection(app.config['db_config'])
        if connection:
            cursor = connection.cursor(dictionary=True)
            cursor.execute(
                'SELECT c.id, c.name, c.is_favorite, GROUP_CONCAT(CONCAT(cm.type, ":", cm.value) SEPARATOR ";") as methods FROM contacts c LEFT JOIN contact_methods cm ON c.id = cm.contact_id GROUP BY c.id')
            contacts = cursor.fetchall()
            cursor.close()
            connection.close()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['ID', '姓名', '是否收藏', '联系方式'])  # 联系方式作为分号分隔的 type:value
            for contact in contacts:
                methods_str = contact['methods'] if contact['methods'] else ''
                ws.append([contact['id'], contact['name'], contact['is_favorite'], methods_str])

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, as_attachment=True, download_name='contacts.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return jsonify({'error': '数据库连接失败'}), 500

    @app.route('/api/import', methods=['POST'])
    def import_contacts():
        if 'file' not in request.files:
            return jsonify({'error': '无文件部分'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400
        if file and file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filename=BytesIO(file.read()))
            ws = wb.active
            connection = get_db_connection(app.config['db_config'])
            if connection:
                cursor = connection.cursor()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    id, name, is_favorite, methods_str = row
                    # 插入联系人（忽略 ID，自动递增）
                    cursor.execute('INSERT INTO contacts (name, is_favorite) VALUES (%s, %s)',
                                   (name, bool(is_favorite)))
                    contact_id = cursor.lastrowid
                    if methods_str:
                        methods = methods_str.split(';')
                        for method in methods:
                            if ':' in method:
                                type_, value = method.split(':', 1)
                                cursor.execute(
                                    'INSERT INTO contact_methods (contact_id, type, value) VALUES (%s, %s, %s)',
                                    (contact_id, type_.strip(), value.strip()))
                connection.commit()
                cursor.close()
                connection.close()
                return jsonify({'message': '联系人已导入'}), 200
            return jsonify({'error': '数据库连接失败'}), 500
        return jsonify({'error': '无效的文件格式'}), 400